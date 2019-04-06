'''
爬取cndn学院课程数据
数据写入.xlsx文件中
'''
import aiohttp
import asyncio
from lxml import etree
from fake_useragent import UserAgent
import openpyxl


def save(datalists):
    book_name_xlsx = 'test.xlsx'
    sheet_name_xlsx = 'xlsx格式test'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name_xlsx
    index = 1
    title = ['href', 'course_lessons', 'title', 'num', 'lecnameellipsis', 'price']
    for i in range(0, len(title)):
        sheet.cell(row=index, column=i + 1, value=title[i])
    index+=1
    for list in datalists:
        sheet.cell(row=index, column=1, value=list[0])
        sheet.cell(row=index, column=2, value=list[1])
        sheet.cell(row=index, column=3, value=list[2])
        sheet.cell(row=index, column=4, value=list[3])
        sheet.cell(row=index, column=5, value=list[4])
        sheet.cell(row=index, column=6, value=list[5])
        index+=1
    workbook.save(book_name_xlsx)


ua=UserAgent()
headers={'User-Agent':ua.random}
datalists=[]
def prase_content(content):
    tree=etree.HTML(content)
    lists=tree.xpath('//div[@class="course_item"]')
    print(len(lists))
    for list in lists:
        datalist=[]
        datalist.append(list.xpath(' ./a[1]/@href')[0])
        datalist.append(list.xpath('string(.//span[@class="course_lessons"])'))
        datalist.append(list.xpath('string(.//span[@class="title ellipsis-2"])').replace('[会员秒杀]','').strip())
        datalist.append(list.xpath('string(.//p[@class="subinfo"]/span[1])').strip())
        datalist.append(list.xpath('string(.//p[@class="subinfo"]/span[last()])'))
        datalist.append(list.xpath('string(.//p[@class="priceinfo clearfix"])').strip().replace('\n','').replace(' ',''))
        datalists.append(datalist)
    print('添加数据成功')

async def get_info(url):
    async with aiohttp.ClientSession(headers=headers) as Session:
        try:
            async with Session.get(url,timeout=5) as response:
                content=await response.text()
                #print(content)
                prase_content(content)
        except Exception as e:
            print(e)

sema=asyncio.Semaphore(3)
async def get(url):
    with await(sema):
        await get_info(url)

def main():
    url='https://edu.csdn.net/courses/p{}'
    urls=[url.format(page) for page in range(1,3)]
    loop=asyncio.get_event_loop()
    tasks=[get(url)for url in urls]
    loop.run_until_complete(asyncio.wait(tasks))
    save(datalists)
if __name__ == '__main__':
    main()